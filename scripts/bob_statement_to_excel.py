#!/usr/bin/env python3
"""Convert a Bank of Baroda savings-account PDF statement (two-line transaction
layout: summary line + narration line) into a clean Excel workbook.

Usage:
    python bob_statement_to_excel.py "BOB 7593.pdf" [output.xlsx]
"""

import re
import sys
from collections import defaultdict
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

DATE_RE = re.compile(r"^\d{2}-\d{2}-\d{2}$")
BALANCE_RE = re.compile(r"^([\d,]+\.\d{2})(Cr|Dr)$", re.IGNORECASE)
PLAIN_AMOUNT_RE = re.compile(r"^[\d,]+(?:\.\d{2})?$")
HEADER_COLUMNS = ["DATE", "PARTICULARS", "CHQ.NO.", "WITHDRAWALS", "DEPOSITS", "BALANCE"]


def parse_amount(text):
    """'1,26,855.44' -> 126855.44 (Indian comma grouping; openpyxl doesn't care
    about grouping, only Python's float() does, hence the comma strip)."""
    if text is None:
        return 0.0
    cleaned = text.replace(",", "").strip()
    if not cleaned or cleaned == "0":
        return 0.0
    return float(cleaned)


def group_words_into_lines(page):
    """pdfplumber gives words, not lines; group by rounded 'top' so wrapped
    glyphs on the same physical line land in one bucket."""
    lines = defaultdict(list)
    for word in page.extract_words(use_text_flow=False, keep_blank_chars=False):
        lines[round(word["top"])].append(word)
    return [sorted(lines[top], key=lambda w: w["x0"]) for top in sorted(lines)]


def find_header_columns(row_words):
    """Match the repeated 'DATE PARTICULARS CHQ.NO. WITHDRAWALS DEPOSITS
    BALANCE' row and return {column_name: x0} to use as bucket boundaries."""
    texts = [w["text"] for w in row_words]
    if texts[:2] != ["DATE", "PARTICULARS"]:
        return None
    lookup = {w["text"]: w["x0"] for w in row_words}
    if not all(col in lookup for col in HEADER_COLUMNS):
        return None
    return {col: lookup[col] for col in HEADER_COLUMNS}


def bucket_column(word, boundaries):
    """Assign a word to the column whose boundary is closest to the word's
    horizontal centre. Centre (not x0) because amounts are right-aligned
    within their column, so a short '75.00' and a long '1,26,855.44Cr' start
    at very different x-offsets even inside the same column."""
    centre = (word["x0"] + word["x1"]) / 2
    names = list(boundaries.keys())
    starts = [boundaries[n] for n in names]
    best = names[0]
    for name, start in zip(names, starts):
        if centre >= start:
            best = name
    return best


def parse_pdf(path):
    transactions = []
    opening_balance = None
    page_totals = []  # (withdrawals, deposits, balance) per page, for validation
    grand_total = None
    clr_bal = None

    # boundaries/pending persist across pages: not every physical page reprints
    # the letterhead+header block -- a transaction list can spill onto the next
    # page with no header at all, straight back into transaction rows.
    boundaries = None
    pending = None  # transaction dict awaiting its narration continuation line

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            rows = group_words_into_lines(page)

            for row_words in rows:
                line_text = " ".join(w["text"] for w in row_words)

                header_cols = find_header_columns(row_words)
                if header_cols:
                    boundaries = header_cols
                    continue

                if line_text.startswith("Page Total:"):
                    amounts = re.findall(r"[\d,]+(?:\.\d{2})?(?:Cr|Dr)?", line_text[len("Page Total:"):])
                    amounts = [a for a in amounts if a]
                    if len(amounts) >= 3:
                        withdrawals = parse_amount(amounts[0])
                        deposits = parse_amount(amounts[1])
                        bal_match = BALANCE_RE.match(amounts[2])
                        balance = parse_amount(bal_match.group(1)) if bal_match else parse_amount(amounts[2])
                        page_totals.append((withdrawals, deposits, balance))
                    if pending:
                        transactions.append(pending)
                        pending = None
                    continue

                if line_text.startswith("Grand Total:") and grand_total is None:
                    amounts = re.findall(r"[\d,]+\.\d{2}(?:Cr|Dr)?", line_text)
                    if len(amounts) >= 3:
                        bal_match = BALANCE_RE.match(amounts[2])
                        grand_total = {
                            "withdrawals": parse_amount(amounts[0]),
                            "deposits": parse_amount(amounts[1]),
                            "balance": parse_amount(bal_match.group(1)) if bal_match else parse_amount(amounts[2]),
                            "balance_indicator": bal_match.group(2) if bal_match else None,
                        }
                    continue

                if line_text.startswith("ClrBal:"):
                    m = re.search(r"ClrBal:\s*([\d,]+\.\d{2})", line_text)
                    if m:
                        clr_bal = parse_amount(m.group(1))
                    continue

                if boundaries is None:
                    continue  # preamble before the first header row seen on page 1

                first_word = row_words[0]["text"]

                if DATE_RE.match(first_word):
                    if pending:
                        transactions.append(pending)

                    cols = defaultdict(list)
                    for w in row_words[1:]:
                        cols[bucket_column(w, boundaries)].append(w["text"])

                    balance_text = " ".join(cols["BALANCE"])
                    bal_match = BALANCE_RE.match(balance_text.replace(" ", ""))
                    balance = parse_amount(bal_match.group(1)) if bal_match else None
                    cr_dr = bal_match.group(2).title() if bal_match else None

                    pending = {
                        "date": first_word,
                        "particulars": " ".join(cols["PARTICULARS"]),
                        "chq_no": " ".join(cols["CHQ.NO."]),
                        "withdrawal": parse_amount(" ".join(cols["WITHDRAWALS"])) if cols["WITHDRAWALS"] else 0.0,
                        "deposit": parse_amount(" ".join(cols["DEPOSITS"])) if cols["DEPOSITS"] else 0.0,
                        "balance": balance,
                        "cr_dr": cr_dr,
                        "narration": "",
                    }

                    if pending["particulars"] == "B/F" and opening_balance is None:
                        opening_balance = pending
                        pending = None
                else:
                    # Narration continuation line belonging to the previous transaction.
                    if pending:
                        pending["narration"] = (pending["narration"] + " " + line_text).strip()

    if pending:
        transactions.append(pending)

    return opening_balance, transactions, page_totals, grand_total, clr_bal


def write_excel(path, opening_balance, transactions, page_totals, grand_total, clr_bal):
    wb = Workbook()

    ws = wb.active
    ws.title = "Transactions"
    headers = ["Date", "Particulars", "Chq No", "Withdrawal", "Deposit", "Balance", "Cr/Dr", "Full Narration"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    if opening_balance:
        ws.append([
            opening_balance["date"], "Opening Balance (B/F)", "",
            None, None, opening_balance["balance"], opening_balance["cr_dr"], "",
        ])

    for t in transactions:
        ws.append([
            t["date"], t["particulars"], t["chq_no"],
            t["withdrawal"] or None, t["deposit"] or None,
            t["balance"], t["cr_dr"], t["narration"],
        ])

    widths = [10, 20, 12, 14, 14, 16, 8, 55]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    summary = wb.create_sheet("Validation")
    summary.append(["Check", "Extracted", "From PDF", "Match"])
    for cell in summary[1]:
        cell.font = Font(bold=True)

    total_withdrawals = round(sum(t["withdrawal"] for t in transactions), 2)
    total_deposits = round(sum(t["deposit"] for t in transactions), 2)
    closing_balance = transactions[-1]["balance"] if transactions else None
    closing_indicator = transactions[-1]["cr_dr"] if transactions else None

    pdf_withdrawals = grand_total["withdrawals"] if grand_total else None
    pdf_deposits = grand_total["deposits"] if grand_total else None
    pdf_closing = grand_total["balance"] if grand_total else None

    def row(label, extracted, pdf_value):
        match = "YES" if pdf_value is not None and abs(extracted - pdf_value) < 0.01 else "NO"
        summary.append([label, extracted, pdf_value, match])

    summary.append(["Transaction count", len(transactions), "", ""])
    row("Sum of Withdrawals", total_withdrawals, pdf_withdrawals)
    row("Sum of Deposits", total_deposits, pdf_deposits)
    row("Closing Balance", closing_balance, pdf_closing)
    summary.append(["Closing indicator", closing_indicator, grand_total["balance_indicator"].title() if grand_total else "", ""])
    summary.append(["ClrBal (as-on, informational only)", "", clr_bal, ""])

    summary.append([])
    summary.append(["Page-by-page totals (withdrawals, deposits, balance)"])
    for i, (w, d, b) in enumerate(page_totals, start=1):
        summary.append([f"Page {i}", w, d, b])

    for col, width in zip("ABCD", [36, 16, 16, 10]):
        summary.column_dimensions[col].width = width

    wb.save(path)
    return total_withdrawals, total_deposits, closing_balance, closing_indicator


def main():
    if len(sys.argv) < 2:
        print("Usage: python bob_statement_to_excel.py <input.pdf> [output.xlsx]")
        sys.exit(1)

    pdf_path = Path(sys.argv[1])
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else pdf_path.with_suffix(".xlsx")

    opening_balance, transactions, page_totals, grand_total, clr_bal = parse_pdf(pdf_path)
    total_withdrawals, total_deposits, closing_balance, closing_indicator = write_excel(
        out_path, opening_balance, transactions, page_totals, grand_total, clr_bal
    )

    print(f"Wrote {out_path}")
    print()
    print("=== Sanity check ===")
    print(f"Opening balance : {opening_balance['balance']:,.2f} {opening_balance['cr_dr']}")
    print(f"Transaction count (excl. opening balance / totals): {len(transactions)}")
    print(f"Sum of Withdrawals : {total_withdrawals:,.2f}")
    print(f"Sum of Deposits    : {total_deposits:,.2f}")
    print(f"Closing balance    : {closing_balance:,.2f} {closing_indicator}")
    if grand_total:
        print()
        print("=== PDF's own Grand Total (for comparison) ===")
        print(f"Grand Total Withdrawals : {grand_total['withdrawals']:,.2f}")
        print(f"Grand Total Deposits    : {grand_total['deposits']:,.2f}")
        print(f"Grand Total Balance     : {grand_total['balance']:,.2f} {grand_total['balance_indicator'].title()}")
        w_ok = abs(total_withdrawals - grand_total["withdrawals"]) < 0.01
        d_ok = abs(total_deposits - grand_total["deposits"]) < 0.01
        b_ok = closing_balance is not None and abs(closing_balance - grand_total["balance"]) < 0.01
        print()
        print(f"Withdrawals match : {'PASS' if w_ok else 'FAIL'}")
        print(f"Deposits match    : {'PASS' if d_ok else 'FAIL'}")
        print(f"Closing bal match : {'PASS' if b_ok else 'FAIL'}")
    if clr_bal is not None:
        print()
        print(f"ClrBal (as-on {'11-08-2026'}): {clr_bal:,.2f} -- this is a live cleared-balance snapshot from the")
        print("bank, not the statement's period-end closing balance, so it is not expected to match Grand Total.")


if __name__ == "__main__":
    main()
